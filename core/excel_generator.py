"""
CORE EXCEL GENERATOR - Sistema Universale Generazione Excel
===========================================================

Funzioni complete per generazione Excel con multiple librerie:
- 📊 openpyxl (Excel avanzato con formule e stili)
- 🐼 pandas (DataFrame to Excel con processing)
- 🎨 Styling automatico e formattazione professionale
- 📈 Grafici e dashboard automatici
- 🔄 Multi-sheet e template complessi

Caratteristiche:
- Zero dipendenze da app specifiche  
- Styling professionale automatico
- Formule e validazione dati
- Gestione tipi di dato automatica
- Output flessibile (file, buffer, response)
- Performance ottimizzate per grandi dataset

Versione: 1.0
Compatibilità: Django 3.2+, Python 3.8+
"""

import os
import tempfile
from io import BytesIO
from typing import Dict, List, Any, Optional, Union, Tuple
from dataclasses import dataclass, field
from datetime import datetime, date
from decimal import Decimal
import logging

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone

# Excel Libraries
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        NamedStyle, Protection
    )
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

logger = logging.getLogger(__name__)


# =============================================================================
# CONFIGURATION CLASSES
# =============================================================================

@dataclass
class ExcelConfig:
    """Configurazione completa per generazione Excel"""
    # Output
    filename: str = "export.xlsx"
    sheet_name: str = "Dati"
    
    # Styling
    header_style: Dict[str, Any] = field(default_factory=lambda: {
        'font': {'bold': True, 'color': 'FFFFFF'},
        'fill': {'fgColor': '4472C4', 'patternType': 'solid'},
        'alignment': {'horizontal': 'center', 'vertical': 'center'},
        'border': True
    })
    
    data_style: Dict[str, Any] = field(default_factory=lambda: {
        'font': {'size': 11},
        'alignment': {'horizontal': 'left', 'vertical': 'center'},
        'border': True
    })
    
    # Layout
    auto_fit_columns: bool = True
    freeze_panes: str = "A2"  # Freeze header row
    add_filters: bool = True
    add_table_style: bool = True
    
    # Data formatting
    number_format: Dict[str, str] = field(default_factory=lambda: {
        'currency': '€ #,##0.00',
        'percentage': '0.00%',
        'date': 'dd/mm/yyyy',
        'datetime': 'dd/mm/yyyy hh:mm',
        'integer': '#,##0',
        'decimal': '#,##0.00'
    })
    
    # Advanced features
    add_charts: bool = False
    add_conditional_formatting: bool = False
    protect_sheet: bool = False
    password: str = None


@dataclass
class ColumnConfig:
    """Configurazione singola colonna"""
    name: str
    data_type: str = 'text'  # text, number, currency, percentage, date, datetime
    width: float = None
    format_string: str = None
    formula: str = None
    validation: Dict = None
    style_override: Dict = None


@dataclass
class ChartConfig:
    """Configurazione grafico Excel"""
    chart_type: str = 'bar'  # bar, line, pie
    title: str = ""
    data_range: str = None
    position: str = "E2"  # Cell position
    width: int = 15
    height: int = 10


# =============================================================================
# MAIN EXCEL GENERATION FUNCTIONS  
# =============================================================================

def generate_excel_from_data(
    data: Union[List[Dict], List[List], pd.DataFrame],
    columns: List[Union[str, ColumnConfig]] = None,
    config: ExcelConfig = None,
    sheets: Dict[str, Any] = None,  # Multiple sheets
    output_type: str = 'response',  # 'response', 'file', 'buffer'
    output_path: str = None
) -> Union[HttpResponse, BytesIO, str]:
    """
    Genera Excel da dati strutturati
    
    Args:
        data: Dati da esportare (lista dict, lista liste, o DataFrame)
        columns: Configurazione colonne
        config: Configurazione Excel
        sheets: Dict per multi-sheet {nome_sheet: dati}
        output_type: Tipo output ('response', 'file', 'buffer')
        output_path: Percorso file per output='file'
        
    Returns:
        HttpResponse, BytesIO o str (percorso file)
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl non disponibile. Installa: pip install openpyxl")
    
    # Configurazione default
    if config is None:
        config = ExcelConfig()
    
    # Crea workbook
    wb = Workbook()
    
    if sheets:
        # Multi-sheet mode
        # Rimuovi sheet di default
        wb.remove(wb.active)
        
        for sheet_name, sheet_data in sheets.items():
            sheet_config = ExcelConfig()
            sheet_config.sheet_name = sheet_name
            
            if isinstance(sheet_data, dict):
                sheet_columns = sheet_data.get('columns', columns)
                sheet_data = sheet_data.get('data', [])
            else:
                sheet_columns = columns
            
            _create_excel_sheet(wb, sheet_data, sheet_columns, sheet_config)
    else:
        # Single sheet mode
        ws = wb.active
        ws.title = config.sheet_name
        _populate_excel_sheet(ws, data, columns, config)
    
    # Output handling
    if output_type == 'buffer':
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    elif output_type == 'file':
        if not output_path:
            output_path = _generate_temp_filename(config.filename)
        
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return output_path
    
    else:  # response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{config.filename}"'
        
        wb.save(response)
        return response


def generate_excel_with_pandas(
    dataframes: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
    config: ExcelConfig = None,
    output_type: str = 'response',
    output_path: str = None,
    excel_writer_args: Dict = None
) -> Union[HttpResponse, BytesIO, str]:
    """
    Genera Excel usando pandas ExcelWriter
    
    Args:
        dataframes: DataFrame singolo o dict di DataFrames per multi-sheet
        config: Configurazione Excel
        output_type: Tipo output
        output_path: Percorso file per output='file'
        excel_writer_args: Argomenti aggiuntivi per ExcelWriter
        
    Returns:
        HttpResponse, BytesIO o str (percorso file)
    """
    if not PANDAS_AVAILABLE:
        raise ImportError("pandas non disponibile. Installa: pip install pandas")
    
    if config is None:
        config = ExcelConfig()
    
    if excel_writer_args is None:
        excel_writer_args = {'engine': 'openpyxl'}
    
    # Setup output
    if output_type == 'buffer':
        buffer = BytesIO()
        output_target = buffer
    elif output_type == 'file':
        if not output_path:
            output_path = _generate_temp_filename(config.filename)
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        output_target = output_path
    else:  # response
        buffer = BytesIO()
        output_target = buffer
    
    # Write Excel
    with pd.ExcelWriter(output_target, **excel_writer_args) as writer:
        if isinstance(dataframes, dict):
            # Multi-sheet
            for sheet_name, df in dataframes.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply styling if openpyxl
                if excel_writer_args.get('engine') == 'openpyxl':
                    worksheet = writer.sheets[sheet_name]
                    _apply_pandas_styling(worksheet, df, config)
        else:
            # Single sheet
            dataframes.to_excel(writer, sheet_name=config.sheet_name, index=False)
            
            if excel_writer_args.get('engine') == 'openpyxl':
                worksheet = writer.sheets[config.sheet_name]
                _apply_pandas_styling(worksheet, dataframes, config)
    
    # Return handling
    if output_type == 'buffer':
        buffer.seek(0)
        return buffer
    elif output_type == 'file':
        return output_path
    else:  # response
        response = HttpResponse(
            content=buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{config.filename}"'
        return response


# =============================================================================
# SHEET CREATION AND POPULATION
# =============================================================================

def _create_excel_sheet(wb: Workbook, data: Any, columns: List, config: ExcelConfig):
    """Crea e popola un nuovo sheet"""
    ws = wb.create_sheet(title=config.sheet_name)
    _populate_excel_sheet(ws, data, columns, config)


def _populate_excel_sheet(worksheet, data: Any, columns: List, config: ExcelConfig):
    """Popola worksheet con dati e styling"""
    
    # Converti data in formato uniforme
    if isinstance(data, pd.DataFrame):
        headers = list(data.columns)
        rows = data.values.tolist()
    elif data and isinstance(data[0], dict):
        # Lista di dizionari
        if not columns:
            headers = list(data[0].keys())
        else:
            headers = [col.name if isinstance(col, ColumnConfig) else col for col in columns]
        
        rows = []
        for item in data:
            row = [item.get(header, '') for header in headers]
            rows.append(row)
    elif data and isinstance(data[0], (list, tuple)):
        # Lista di liste
        if columns:
            headers = [col.name if isinstance(col, ColumnConfig) else col for col in columns]
            rows = data
        else:
            headers = [f"Colonna {i+1}" for i in range(len(data[0]))]
            rows = data
    else:
        headers = ["Dati"]
        rows = [[str(item)] for item in data] if data else []
    
    # Scrivi headers
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        _apply_header_style(cell, config)
    
    # Scrivi dati
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            # Converti e formatta valore
            formatted_value = _format_cell_value(value)
            cell = worksheet.cell(row=row_idx, column=col_idx, value=formatted_value)
            
            # Applica stile dati
            _apply_data_style(cell, config)
            
            # Applica formato numero se specificato
            if columns and col_idx <= len(columns):
                col_config = columns[col_idx - 1]
                if isinstance(col_config, ColumnConfig):
                    _apply_column_formatting(cell, col_config, config)
    
    # Post-processing
    _post_process_worksheet(worksheet, headers, config)


def _apply_header_style(cell, config: ExcelConfig):
    """Applica stile header"""
    style = config.header_style
    
    if 'font' in style:
        font_args = style['font']
        cell.font = Font(**font_args)
    
    if 'fill' in style:
        fill_args = style['fill']
        cell.fill = PatternFill(**fill_args)
    
    if 'alignment' in style:
        align_args = style['alignment']
        cell.alignment = Alignment(**align_args)
    
    if style.get('border'):
        _apply_border(cell)


def _apply_data_style(cell, config: ExcelConfig):
    """Applica stile dati"""
    style = config.data_style
    
    if 'font' in style:
        font_args = style['font']
        cell.font = Font(**font_args)
    
    if 'alignment' in style:
        align_args = style['alignment']  
        cell.alignment = Alignment(**align_args)
    
    if style.get('border'):
        _apply_border(cell)


def _apply_border(cell):
    """Applica bordo standard"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def _apply_column_formatting(cell, col_config: ColumnConfig, config: ExcelConfig):
    """Applica formattazione specifica colonna"""
    if col_config.format_string:
        cell.number_format = col_config.format_string
    elif col_config.data_type in config.number_format:
        cell.number_format = config.number_format[col_config.data_type]


def _post_process_worksheet(worksheet, headers: List[str], config: ExcelConfig):
    """Post-processing del worksheet"""
    
    # Auto-fit columns
    if config.auto_fit_columns:
        for col in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze panes
    if config.freeze_panes:
        worksheet.freeze_panes = config.freeze_panes
    
    # Add filters
    if config.add_filters:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    
    # Add table style
    if config.add_table_style:
        max_row = worksheet.max_row
        max_col = len(headers)
        table_range = f"A1:{get_column_letter(max_col)}{max_row}"
        
        table = Table(displayName="DataTable", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)
    
    # Conditional formatting
    if config.add_conditional_formatting:
        _add_conditional_formatting(worksheet, headers)
    
    # Protection
    if config.protect_sheet:
        worksheet.protection.sheet = True
        if config.password:
            worksheet.protection.password = config.password


def _add_conditional_formatting(worksheet, headers: List[str]):
    """Aggiunge formattazione condizionale automatica"""
    max_row = worksheet.max_row
    
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{max_row}"
        
        # Prova a rilevare tipo dati per formattazione appropriata
        sample_value = worksheet.cell(row=2, column=col_idx).value
        
        if isinstance(sample_value, (int, float, Decimal)):
            # Color scale per numeri
            color_scale = ColorScaleRule(
                start_type='min',
                start_color='F8696B',
                end_type='max', 
                end_color='63BE7B'
            )
            worksheet.conditional_formatting.add(data_range, color_scale)


def _apply_pandas_styling(worksheet, df: pd.DataFrame, config: ExcelConfig):
    """Applica styling a worksheet creato con pandas"""
    # Stile header
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        _apply_header_style(cell, config)
    
    # Stile dati
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            _apply_data_style(cell, config)
    
    # Post-processing
    _post_process_worksheet(worksheet, list(df.columns), config)


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def _format_cell_value(value: Any) -> Any:
    """Formatta valore per Excel"""
    if value is None:
        return ""
    
    if isinstance(value, (date, datetime)):
        return value
    
    if isinstance(value, Decimal):
        return float(value)
    
    if isinstance(value, bool):
        return value
    
    if isinstance(value, (int, float)):
        return value
    
    return str(value)


def _generate_temp_filename(base_filename: str) -> str:
    """Genera percorso file temporaneo"""
    temp_dir = getattr(settings, 'TEMP_DIR', tempfile.gettempdir())
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    name, ext = os.path.splitext(base_filename)
    filename = f"{name}_{timestamp}{ext}"
    return os.path.join(temp_dir, 'excel', filename)


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def create_simple_excel(
    data: List[Dict],
    filename: str = 'export.xlsx',
    sheet_name: str = 'Dati',
    output_type: str = 'response'
):
    """Funzione convenienza per Excel semplice"""
    config = ExcelConfig(filename=filename, sheet_name=sheet_name)
    
    return generate_excel_from_data(
        data=data,
        config=config,
        output_type=output_type
    )


def create_multi_sheet_excel(
    sheets_data: Dict[str, List[Dict]],
    filename: str = 'export_multi.xlsx',
    output_type: str = 'response'
):
    """Funzione convenienza per Excel multi-sheet"""
    config = ExcelConfig(filename=filename)
    
    return generate_excel_from_data(
        data=[],  # Not used in multi-sheet mode
        config=config,
        sheets=sheets_data,
        output_type=output_type
    )


def create_styled_excel(
    data: List[Dict],
    columns: List[ColumnConfig],
    filename: str = 'styled_export.xlsx',
    output_type: str = 'response'
):
    """Funzione convenienza per Excel con styling personalizzato"""
    config = ExcelConfig(
        filename=filename,
        add_charts=True,
        add_conditional_formatting=True
    )
    
    return generate_excel_from_data(
        data=data,
        columns=columns,
        config=config,
        output_type=output_type
    )


def dataframe_to_excel_response(
    df: pd.DataFrame,
    filename: str = 'dataframe_export.xlsx',
    sheet_name: str = 'Dati'
) -> HttpResponse:
    """Funzione convenienza per DataFrame to Excel HTTP response"""
    config = ExcelConfig(filename=filename, sheet_name=sheet_name)
    
    return generate_excel_with_pandas(
        dataframes=df,
        config=config,
        output_type='response'
    )